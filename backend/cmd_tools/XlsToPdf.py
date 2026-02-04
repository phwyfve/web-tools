"""
XlsToPdf command handler
Converts Excel files (.xls, .xlsx) to PDF format
"""

import io
import logging
from typing import Dict, Any, List
from bson import ObjectId
import gridfs
import openpyxl
from openpyxl.utils import get_column_letter
from PyPDF2 import PdfWriter, PdfReader
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER

# Set up logging for this module
logger = logging.getLogger('XlsToPdf')

# Supported Excel formats
SUPPORTED_EXCEL_FORMATS = {'.xls', '.xlsx', '.xlsm'}

def xls_to_pdf(args: Dict[str, Any], db, fs: gridfs.GridFS) -> Dict[str, Any]:
    """
    Convert multiple Excel files from GridFS to PDF and merge into a single PDF
    
    Args:
        args: Dict containing 'file_ids' - list of GridFS file IDs
        db: MongoDB database connection
        fs: GridFS instance for tmp_files bucket
        
    Returns:
        Dict containing 'merged_file_id' of the resulting merged PDF
    """
    
    file_ids = args.get("file_ids", [])
    if not file_ids:
        raise ValueError("No file_ids provided for Excel to PDF conversion")
    
    logger.info(f"Starting Excel to PDF conversion of {len(file_ids)} files: {file_ids}")
    
    # Create PDF writer for merged output
    pdf_writer = PdfWriter()
    processed_files = []
    total_sheets_converted = 0
    
    try:
        # Process each Excel file
        for i, file_id_str in enumerate(file_ids):
            logger.info(f"Processing file {i+1}/{len(file_ids)}: {file_id_str}")
            
            # Convert string ID to ObjectId
            try:
                file_id = ObjectId(file_id_str)
            except Exception as e:
                raise ValueError(f"Invalid file ID '{file_id_str}': {str(e)}")
            
            # Download file from GridFS
            try:
                grid_file = fs.get(file_id)
                file_content = grid_file.read()
                processed_files.append({
                    "id": file_id_str,
                    "filename": grid_file.filename,
                    "size": len(file_content)
                })
                logger.info(f"Downloaded {grid_file.filename} ({len(file_content)} bytes)")
            except gridfs.NoFile:
                raise ValueError(f"File with ID '{file_id_str}' not found in GridFS")
            except Exception as e:
                raise ValueError(f"Failed to download file '{file_id_str}': {str(e)}")
            
            # Convert Excel to PDF pages
            try:
                # Load Excel file using openpyxl
                excel_buffer = io.BytesIO(file_content)
                workbook = openpyxl.load_workbook(excel_buffer, data_only=True)
                
                logger.info(f"Excel file loaded. Sheets: {workbook.sheetnames}")
                
                # Process each sheet
                for sheet_index, sheet_name in enumerate(workbook.sheetnames):
                    logger.info(f"Processing sheet {sheet_index + 1}/{len(workbook.sheetnames)}: {sheet_name}")
                    
                    sheet = workbook[sheet_name]
                    
                    # Skip empty sheets
                    if sheet.max_row == 0 or sheet.max_column == 0:
                        logger.info(f"Skipping empty sheet: {sheet_name}")
                        continue
                    
                    # Extract data from sheet
                    data = []
                    for row in sheet.iter_rows(values_only=True):
                        # Convert None values to empty strings and handle all types
                        cleaned_row = []
                        for cell_value in row:
                            if cell_value is None:
                                cleaned_row.append('')
                            else:
                                cleaned_row.append(str(cell_value))
                        data.append(cleaned_row)
                    
                    if not data:
                        logger.info(f"No data in sheet: {sheet_name}")
                        continue
                    
                    logger.info(f"Extracted {len(data)} rows from sheet {sheet_name}")
                    
                    # Create PDF for this sheet
                    pdf_buffer = io.BytesIO()
                    
                    # Determine page orientation based on number of columns
                    num_columns = len(data[0]) if data else 0
                    if num_columns > 8:
                        pagesize = landscape(A4)
                    else:
                        pagesize = A4
                    
                    doc = SimpleDocTemplate(
                        pdf_buffer,
                        pagesize=pagesize,
                        leftMargin=0.5*inch,
                        rightMargin=0.5*inch,
                        topMargin=0.75*inch,
                        bottomMargin=0.5*inch
                    )
                    
                    # Build PDF content
                    elements = []
                    styles = getSampleStyleSheet()
                    
                    # Add sheet title
                    title_style = ParagraphStyle(
                        'SheetTitle',
                        parent=styles['Heading1'],
                        fontSize=14,
                        textColor=colors.HexColor('#1a56db'),
                        spaceAfter=12,
                        alignment=TA_CENTER
                    )
                    
                    title_text = f"{grid_file.filename} - {sheet_name}"
                    elements.append(Paragraph(title_text, title_style))
                    elements.append(Spacer(1, 0.2*inch))
                    
                    # Calculate column widths dynamically
                    page_width = pagesize[0] - doc.leftMargin - doc.rightMargin
                    col_width = page_width / num_columns if num_columns > 0 else page_width
                    
                    # Limit column width for readability
                    max_col_width = 2.5 * inch
                    min_col_width = 0.5 * inch
                    col_width = max(min_col_width, min(col_width, max_col_width))
                    
                    col_widths = [col_width] * num_columns
                    
                    # Create table
                    table = Table(data, colWidths=col_widths, repeatRows=1)
                    
                    # Style the table
                    table_style = TableStyle([
                        # Header row styling
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a56db')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('TOPPADDING', (0, 0), (-1, 0), 12),
                        
                        # Body styling
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 8),
                        ('TOPPADDING', (0, 1), (-1, -1), 6),
                        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                        ('LEFTPADDING', (0, 0), (-1, -1), 6),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                        
                        # Grid
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        
                        # Alternating row colors
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')])
                    ])
                    
                    table.setStyle(table_style)
                    elements.append(table)
                    
                    # Add page break if not the last sheet
                    if sheet_index < len(workbook.sheetnames) - 1:
                        elements.append(PageBreak())
                    
                    # Build PDF
                    doc.build(elements)
                    
                    # Read the created PDF
                    pdf_buffer.seek(0)
                    sheet_pdf = PdfReader(pdf_buffer)
                    
                    # Add all pages from this sheet to the merged PDF
                    for page in sheet_pdf.pages:
                        pdf_writer.add_page(page)
                    
                    total_sheets_converted += 1
                    logger.info(f"Successfully converted sheet {sheet_name} to PDF ({len(sheet_pdf.pages)} pages)")
                
                workbook.close()
                logger.info(f"Completed processing {grid_file.filename}")
                
            except Exception as e:
                logger.error(f"Failed to convert Excel file '{grid_file.filename}': {str(e)}")
                logger.exception("Full exception details:")
                raise ValueError(f"Failed to convert Excel file '{grid_file.filename}' to PDF: {str(e)}")
        
        # Create merged PDF in memory
        merged_buffer = io.BytesIO()
        pdf_writer.write(merged_buffer)
        merged_content = merged_buffer.getvalue()
        
        logger.info(f"Merged PDF created successfully ({len(merged_content)} bytes, {total_sheets_converted} sheets)")
        
        # Generate filename for merged PDF
        if len(file_ids) == 1:
            merged_filename = f"{processed_files[0]['filename'].rsplit('.', 1)[0]}_converted.pdf"
        else:
            merged_filename = f"excel_to_pdf_{len(file_ids)}_files.pdf"
        
        # Upload merged PDF to GridFS
        merged_file_id = fs.put(
            merged_content,
            filename=merged_filename,
            content_type="application/pdf",
            metadata={
                "original_files": processed_files,
                "conversion_type": "excel_to_pdf",
                "total_pages": len(pdf_writer.pages),
                "total_sheets_converted": total_sheets_converted
            }
        )
        
        logger.info(f"Merged PDF uploaded to GridFS with ID: {merged_file_id}")
        
        # Return success result
        return {
            "success": True,
            "merged_file_id": str(merged_file_id),
            "merged_filename": merged_filename,
            "total_pages": len(pdf_writer.pages),
            "total_sheets_converted": total_sheets_converted,
            "original_files": processed_files,
            "merged_size_bytes": len(merged_content)
        }
        
    except Exception as e:
        # Ensure proper error handling
        logger.error(f"Excel to PDF conversion failed: {str(e)}")
        logger.exception("Full exception details:")
        raise  # Re-raise to be caught by myshell.py
